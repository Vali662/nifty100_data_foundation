import sqlite3
import pandas as pd
import yaml
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------
# Load Database
# -----------------------------
conn = sqlite3.connect("data/nifty100.db")

query = """
SELECT
    fr.*,
    mc.pe_ratio,
    mc.pb_ratio,
    mc.dividend_yield_pct,
    mc.market_cap_crore,
    pl.sales,
    pl.net_profit,
    s.broad_sector

FROM financial_ratios fr

LEFT JOIN market_cap mc
ON fr.company_id = mc.company_id
AND CAST(substr(fr.year, -4) AS INTEGER) = mc.year

LEFT JOIN profitandloss pl
ON fr.company_id = pl.company_id
AND fr.year = pl.year

LEFT JOIN sectors s
ON fr.company_id = s.company_id
"""

df = pd.read_sql(query, conn)
df["old_composite_score"] = df["composite_quality_score"]
# Keep only the latest year for each company
df = (
    df.sort_values("year")
      .groupby("company_id", as_index=False)
      .tail(1)
)

print("Companies:", len(df))
print("\nMissing Values:")
print(df[[
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct"
]].isnull().sum())

print("\nStatistics:")
print(df[[
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct"
]].describe())

# -----------------------------
# Load YAML Config
# -----------------------------
with open("config/screener_config.yaml", "r") as file:
    config = yaml.safe_load(file)

# -----------------------------
# Clean Interest Coverage
# -----------------------------
df["interest_coverage"] = (
    pd.to_numeric(df["interest_coverage"], errors="coerce")
      .fillna(float("inf"))
)

def filter_roe(df, value):
    return df[df["return_on_equity_pct"] >= value]


def filter_de(df, value):

    # For Debt-Free preset, require exact zero debt
    if value == 0:
        return df[df["debt_to_equity"] == 0]

    # Financial companies are exempt only for other D/E filters
    financials = df["broad_sector"] == "Financials"

    return df[
        financials |
        (df["debt_to_equity"] <= value)
    ]

def filter_fcf(df, value):
    return df[df["free_cash_flow"] >= value]


def filter_revenue_growth(df, value):
    return df[df["revenue_cagr_5yr"] >= value]


def filter_pat_growth(df, value):
    return df[df["pat_cagr_5yr"] >= value]


def filter_opm(df, value):
    return df[df["operating_profit_margin_pct"] >= value]


def filter_pe(df, value):
    return df[df["pe_ratio"] <= value]


def filter_pb(df, value):
    return df[df["pb_ratio"] <= value]


def filter_dividend(df, value):
    return df[df["dividend_yield_pct"] >= value]


def filter_icr(df, value):
    return df[df["interest_coverage"] >= value]


def filter_market_cap(df, value):
    return df[df["market_cap_crore"] >= value]


def filter_net_profit(df, value):
    return df[df["net_profit"] >= value]


def filter_eps_growth(df, value):
    return df[df["eps_cagr_5yr"] >= value]


def filter_asset_turnover(df, value):
    return df[df["asset_turnover"] >= value]


def filter_sales(df, value):
    return df[df["sales"] >= value]

def filter_dividend_payout(df, value):
    return df[df["dividend_payout_ratio_pct"] <= value]

def apply_filters(df, filters):


    if "roe_min" in filters:
        df = filter_roe(df, filters["roe_min"])

    if "debt_equity_max" in filters:
        df = filter_de(df, filters["debt_equity_max"])

    if "free_cash_flow_min" in filters:
        df = filter_fcf(df, filters["free_cash_flow_min"])

    if "revenue_cagr_5yr_min" in filters:
        df = filter_revenue_growth(df, filters["revenue_cagr_5yr_min"])

    if "pat_cagr_5yr_min" in filters:
        df = filter_pat_growth(df, filters["pat_cagr_5yr_min"])

    if "opm_min" in filters:
        df = filter_opm(df, filters["opm_min"])

    if "pe_max" in filters:
        df = filter_pe(df, filters["pe_max"])

    if "pb_max" in filters:
        df = filter_pb(df, filters["pb_max"])

    if "dividend_yield_min" in filters:
        df = filter_dividend(df, filters["dividend_yield_min"])

    if "interest_coverage_min" in filters:
        df = filter_icr(df, filters["interest_coverage_min"])

    if "market_cap_min" in filters:
        df = filter_market_cap(df, filters["market_cap_min"])

    if "net_profit_min" in filters:
        df = filter_net_profit(df, filters["net_profit_min"])

    if "eps_cagr_5yr_min" in filters:
        df = filter_eps_growth(df, filters["eps_cagr_5yr_min"])

    if "asset_turnover_min" in filters:
        df = filter_asset_turnover(df, filters["asset_turnover_min"])

    if "dividend_payout_max" in filters:
        df = filter_dividend_payout(
        df,
        filters["dividend_payout_max"]
    )

    if "sales_min" in filters:
        df = filter_sales(df, filters["sales_min"])

    return df.sort_values(
        "composite_quality_score",
        ascending=False
    )
import numpy as np

def normalize(series):
    p10 = series.quantile(0.10)
    p90 = series.quantile(0.90)

    series = series.clip(lower=p10, upper=p90)

    minimum = series.min()
    maximum = series.max()

    if maximum == minimum:
        return pd.Series(50, index=series.index)

    return ((series - minimum) / (maximum - minimum)) * 100


def calculate_composite_score(df):

    roe = normalize(df["return_on_equity_pct"])
    roce = normalize(df["return_on_capital_employed_pct"])
    npm = normalize(df["net_profit_margin_pct"])

    revenue = normalize(df["revenue_cagr_5yr"])
    pat = normalize(df["pat_cagr_5yr"])

    fcf = normalize(df["free_cash_flow"])
    cfo = normalize(df["cash_from_operations_cr"])

    debt = 100 - normalize(df["debt_to_equity"])
    icr = normalize(df["interest_coverage"])

    profitability = (
        roe * 0.15 +
        roce * 0.10 +
        npm * 0.10
    )

    cash_quality = (
        fcf * 0.15 +
        cfo * 0.10 +
        (df["free_cash_flow"] > 0).astype(int) * 5
    )

    growth = (
        revenue * 0.10 +
        pat * 0.10
    )

    leverage = (
        debt * 0.10 +
        icr * 0.05
    )
    df["composite_quality_score"] = (
        profitability +
        cash_quality +
        growth +
        leverage
    ).round(2)

    return df

def run_preset(preset_name):

    if preset_name not in config:
        raise ValueError(f"{preset_name} not found")

    result = apply_filters(df, config[preset_name])

    print("\n" + "=" * 60)
    print(f"Preset: {preset_name}")
    print(f"Companies Found: {len(result)}")

    print(result[[
        "company_id",
        "year",
        "return_on_equity_pct",
        "debt_to_equity",
        "composite_quality_score"
    ]].head(10))

    return result

def export_to_excel():

    with pd.ExcelWriter(
        "output/screener_output.xlsx",
        engine="openpyxl"
    ) as writer:

        presets = [
            "quality_compounder",
            "value_pick",
            "growth_accelerator",
            "dividend_champion",
            "debt_free_bluechip",
            "turnaround_watch"
        ]

        for preset in presets:

            result = apply_filters(
                df,
                config[preset]
            )

            result.to_excel(
                writer,
                sheet_name=preset[:31],
                index=False
            )

    print("\nExcel file created successfully!")
    
def format_excel():

    wb = load_workbook("output/screener_output.xlsx")

    # Colors
    green = PatternFill(
        fill_type="solid",
        start_color="90EE90",
        end_color="90EE90"
    )

    red = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    # Loop through every sheet
    for sheet in wb.sheetnames:

        ws = wb[sheet]

        # -----------------------------
        # Header Formatting
        # -----------------------------
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Freeze Header
        ws.freeze_panes = "A2"

        # Store Column Numbers
        headers = {}

        for cell in ws[1]:
            headers[cell.value] = cell.column

        # -----------------------------
        # Color Formatting
        # -----------------------------
        for row in range(2, ws.max_row + 1):

            # ROE
            if "return_on_equity_pct" in headers:

                col = headers["return_on_equity_pct"]

                value = ws.cell(row, col).value

                if value is not None:

                    if value >= 15:
                        ws.cell(row, col).fill = green
                    else:
                        ws.cell(row, col).fill = red

            # Debt to Equity

            if "debt_to_equity" in headers:

                col = headers["debt_to_equity"]

                value = ws.cell(row, col).value

                if value is not None:

                    if value <= 1:
                        ws.cell(row, col).fill = green
                    else:
                        ws.cell(row, col).fill = red

        # -----------------------------
        # Auto Fit Columns
        # -----------------------------
        for column_cells in ws.columns:

            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )

            column_letter = get_column_letter(column_cells[0].column)

            ws.column_dimensions[column_letter].width = length + 3

    wb.save("output/screener_output.xlsx")

    print("Excel formatting completed successfully!")

if __name__ == "__main__":

    df = calculate_composite_score(df)

    presets = [
        "quality_compounder",
        "value_pick",
        "growth_accelerator",
        "dividend_champion",
        "debt_free_bluechip",
        "turnaround_watch"
    ]

    for preset in presets:
        run_preset(preset)

    export_to_excel()
    format_excel()