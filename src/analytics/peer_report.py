import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
conn = sqlite3.connect("data/nifty100.db")
peer_df = pd.read_sql("""
SELECT *
FROM peer_percentiles
""", conn)
peer_group_df = pd.read_sql("""
SELECT
company_id,
peer_group_name,
is_benchmark
FROM peer_groups
""", conn)
report_df = peer_df.merge(
    peer_group_df,
    on=[
        "company_id",
        "peer_group_name"
    ],
    how="left"
)
print(report_df.head())

print()

print(report_df["peer_group_name"].value_counts())

def export_peer_report():

    with pd.ExcelWriter(
        "output/peer_comparison.xlsx",
        engine="openpyxl"
    ) as writer:

        for peer_name in sorted(report_df["peer_group_name"].dropna().unique()):

            group = report_df[
                report_df["peer_group_name"] == peer_name
            ]

            # Convert metrics into columns
            sheet = group.pivot_table(
                index=["company_id", "is_benchmark"],
                columns="metric",
                values="percentile_rank"
            ).reset_index()

            sheet.to_excel(
                writer,
                sheet_name=peer_name[:31],
                index=False
            )
def format_peer_report():

    wb = load_workbook("output/peer_comparison.xlsx")

    green = PatternFill(
        fill_type="solid",
        start_color="90EE90",
        end_color="90EE90"
    )

    yellow = PatternFill(
        fill_type="solid",
        start_color="FFF59D",
        end_color="FFF59D"
    )

    red = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    gold = PatternFill(
        fill_type="solid",
        start_color="FFD966",
        end_color="FFD966"
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for ws in wb.worksheets:

        # Header formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Data formatting
        for row in range(2, ws.max_row + 1):

            # Benchmark row
            benchmark = ws.cell(row, 2).value

            if benchmark == 1:
                for cell in ws[row]:
                    cell.fill = gold

            # Percentile columns start after company_id and is_benchmark
            for col in range(3, ws.max_column + 1):

                value = ws.cell(row, col).value

                if isinstance(value, (int, float)):

                    if value >= 75:
                        ws.cell(row, col).fill = green

                    elif value >= 25:
                        ws.cell(row, col).fill = yellow

                    else:
                        ws.cell(row, col).fill = red

    wb.save("output/peer_comparison.xlsx")

    print("Peer comparison formatting completed.")

    print("Peer comparison Excel created successfully!")

if __name__ == "__main__":

    export_peer_report()

    format_peer_report()